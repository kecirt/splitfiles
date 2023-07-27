import openpyxl
import sys
import os
import pandas as pd
from pathlib import Path
from time import strftime
from termcolor import cprint
from draft_emails import get_file_names


def create_workbook_folder(folder):
    """
    Create a folder for the workbooks.
    """
    folder_path = os.path.abspath(folder)
    folder_path = Path(folder_path).parent
    new_folder_name = strftime("%Y-%m-%d") + "_workbooks"
    new_folder_path = os.path.join(folder_path, new_folder_name)
    Path(new_folder_path).mkdir(parents=True, exist_ok=True)
    return new_folder_path


def determine_service_providers(folder):
    """
    Determine service providers in a folder
    """
    files = get_file_names(folder)
    service_providers = []
    for file in files:
        try:
            service_provider = file.split("-")[4]
            service_providers.append(service_provider.split(".")[0])
        except IndexError:
            pass

    service_providers = set(service_providers)
    if("safaricom(1)" in service_providers):
        service_providers.remove("safaricom(1)")
    if("safaricom(2)" in service_providers):
        service_providers.remove("safaricom(2)")
    service_providers = list(service_providers)
    service_providers.append("safaricom")

    return service_providers


def determine_advisory_name(folder):
    """
    Determine advisories
    """
    files = get_file_names(folder)
    if files:
        advisory = files[0].split("-")[3]
        return advisory
    else:
        return None


def create_one_workbook(service_provider, folder):
    """
    Create one workbook
    """
    parent_directory = Path(os.path.abspath(folder))
    file_name = strftime("%Y-%m-%d") + "-" + service_provider + ".xlsx"
    final_path = parent_directory / file_name
    if(not os.path.exists(final_path)):
        new_wb = openpyxl.Workbook()
        try:
            new_wb.save(final_path)
        except PermissionError:
            cprint("[!!!] Error: The file seems to be open.", "red")
            exit(0)

    return final_path


def create_service_workbook(service_providers, folder):
    """
    Create workbooks for each service provider
    """
    files = []
    for service_provider in service_providers:
        path = create_one_workbook(service_provider, folder)
        files.append(path)
    return files


def write_file_to_worksheet(source_file, dest_file, advisory_name):
    """
    Copy csv contents to worksheet
    """
    source_content = pd.read_csv(source_file)
    sheet_title = advisory_name.title()
    with pd.ExcelWriter(dest_file, engine = "openpyxl", mode="a", if_sheet_exists="replace") as writer:
        source_content.to_excel(writer, sheet_title, index=False)
    wkbook = openpyxl.load_workbook(dest_file)
    if("Sheet" in wkbook.sheetnames):
        sheet = wkbook.get_sheet_by_name("Sheet")
        wkbook.remove(sheet)
        wkbook.save(dest_file)


def write_files(folder):
    """
    Write multiple
    """
    workbooks_folder = create_workbook_folder(folder)
    sps = determine_service_providers(folder)
    advisory_name = determine_advisory_name(folder)
    if (advisory_name is not None):
        workbooks = create_service_workbook(sps, workbooks_folder)
        file_names = get_file_names(folder)
        for workbook in workbooks:
            service_provider = os.path.basename(workbook).split("-")[3].split(".")[0]
            for data_file in file_names:
                parent_dir = Path(os.path.abspath(folder))
                file_path = parent_dir / data_file
                if(service_provider in os.path.basename(data_file)):
                    try:
                        write_file_to_worksheet(file_path, workbook, advisory_name)
                    except PermissionError:
                        cprint("[!!!] Error: Failed on: " + str(workbook) + "! File might be open or user doesn't have permission to overwrite.", "red")
                        exit()


def main():
    if(len(sys.argv) >= 2):
        for folder in sys.argv[1:]:
            write_files(folder)
    cprint("[!] INFO: Success!", "green")


if __name__ == "__main__":
    main()
