import openpyxl
import sys
import os.path
import pandas as pd
from pathlib import Path
from time import strftime
from termcolor import cprint
from titlecase import titlecase
from writeemails import get_file_names, get_countries_in_folder
from draft_emails import isp_dict, get_isps_in_folder, get_const_isp

descriptions_file = "advisories_descriptions.xlsx"
directory_path = os.path.dirname(os.path.realpath(__file__))
descriptions_path = os.path.join(directory_path, descriptions_file)
status = "Technical advisory sent to the service provider"

# Dictionary of filename keys to Advisory names
advisory_dict = {
    "andromeda": "Andromeda", "necurs": "Necurs", "pykspa": "Pykspa", "emotet" : "Emotet",
    "bedep": "Bedep", "nymaim": "Nymaim", "m0yv": "M0yv", "gamarue": "Gamarue",
     "tinba": "Tinba", "qsnatch": "QSnatch", "zeus" : "Zeus","matsnu" : "Matsnu", "ranbyus" : "Ranbyus",
    "isakmp": "ISAKMP", "ssdp": "SSDP",
    "brute_force": "Brute Force (Shadowserver)",
    "telnet": "Telnet", "portmapper": "Portmapper",
    "scan_poodle": "SSLv3/Poodle", "memcached": "Memcached",
    "ms-sql": "MS-SQL", "ldap": "LDAP UDP", "smb": "SMB",
    "scan_freak": "SSL/TLS (Freak)", "blacklist": "Blacklist (IntelMQ)",
    "blacklist-kenya": "Blacklisted Ips", "blocklist-kenya": "Blocklist",
    "botnet_drone": "Botnet Drone", "mdns": "mDNS",
    "mongodb": "MongoDB", "netbios": "NetBIOS", "rdp": "RDP",
    "snmp": "SNMP", "tftp": "Open TFTP", "nat_pmp": "NAT-PMP",
    "scan_dns": "DNS Open Resolver", "scan_cwmp": "Open CWMP",
    "scan_vnc": "Accessible VNC Service",
    "scan_ntp-": "NTP-version", "scan_ntpmonitor": "NTP-monitor",
    "scan_ftp": "Accessible FTP",
    "scan_adb": "Accessible Android Debug Bridge",
    "scan_elasticsearch": "Elasticsearch", "darknet": "Darknet",
    "microsoft_sinkhole": "Microsoft Sinkhole",
    "ddos_amplification": "Amplification DDoS Victim",
    "scan_rsync": "Accessible rsync Service",
    "cisco_smart": "Accessible Cisco Smart Install",
    "scan_xdmcp": "Accessible XDMCP Service", "scan_ipmi": "IPMI",
    "scan_qotd": "QOTD",
    "hp_http_scan": "HTTP Scanners", "scan_ics": "ICS Scanners",
    "ubiquiti": "Open Ubiquiti", "chargen": "Open Chargen",
    "compromised_website": "Compromised Website",
    "phishing": "Phishing (IntelMQ)", "http_vulnerable": "Accessible HTTP", "web-": "Web Application (Hornet)",
    "bruteforce-": "Brute Force (Hornet)", "malware": "Malware (Hornet)",
    "ransomware": "Ransomware", "scan_redis": "Open Redis", "scan_ard": "Apple Remote Desktop",
    "scan_afp": "Apple Filing Protocol", "db2": "Open DB2", "ldap_tcp": "LDAP TCP",
    "intel_ssh": "SSH (IntelMQ)", "intel_darknet": "Darknet (IntelMQ)",
    "suppobox": "Suppobox","ipp": "Open IPP","scan_coap": "Accessible CoAP",
    "scan_mqtt": "Open MQTT", "scan_hadoop": "Accessible Hadoop", "scan_radmin": "Exposed Radmin Service",
    "scan_exchange": "Vulnerable Exchange Server", "scan_rdpeudp": "Accessible ms-rdpeudp",
    "event4_sinkhole": "Sinkhole Events","event6_sinkhole": "Sinkholev6 Events", "honeypot_brute_force": "Honeypot Brute Force",
    "honeypot_ddos": "Honeypot DDoS", "honeypot_http_scan": "Honeypot HTTP Scanners", "honeypot_darknet": "Darknet Events", "scan_smtp": "SMTP",
    "microsoft_events": "Microsoft Sinkhole HTTP", "bruteforce_kenya": "Brute Force (Hornet)", "malware_kenya" : "Malware (Hornet)", "webapp_kenya" : "Web Application (Hornet)",
    "bruteforce_global": "Brute Force (Global)","malware_global" : "Malware (Global)", "webapp_global" : "Web Application (Global)",
    "sandbox_url": "Sandbox URL", "device_id": "Device IDv4", "ip_spoofer": "IP Spoofer", "vulnerable_log4j_server": "Vulnerable Log4j Server", "scan6_telnet": "Telnet v6",
    "id6": "Device IDv6", "scan_amqp": "AMQP","adb_kenya" : "Android debug bridge (Hornet)", "adb_global" : "Android debug bridge (Global)", "cisco_global": "Cisco",
    "elastic_global": "Elastic", "idsevents_global": "IDs Events", "idsevents_kenya": "IDs Events", "loginattempts_global": "Login Attempts",
    "tcpudp_global": "TCU/UDP Attacks", "tcpudp_kenya": "TCU/UDP Attacks", "scan6_ssl": "Accessible SSL IPv6", "scan6_ssh": "Accessible SSH IPv6", "scan6_smtp": "Accessible SMTP IPv6", "scan_http6": "Accessible HTTP IPv6",
    "scan_ssl": "Accessible SSL IPv4", "scan_ssh": "Accessible SSH IPv4", "sinkhole_dns": "Sinkhole DNS Events", "scan_quic": "Accessible QUIC", "scan_mysql": "Accessible MySQL IPv4",
    "scan6_mysql": "Accessible MySQL IPv6", "scan_kubernetes": "Accessible Kubernetes", "scan_epmd": "Accessible Erlang Port", "scan_dvr": "Open DVR DHCPDiscover",
    "scan_ddos": "Vulnerable DDoS Middlebox", "scan_socks": "Accessible SOCKS", "rdp_global": "RDP", "rdp_kenya": "RDP",
    "scan_http4": "Accessible HTTP IPv4", "scan_ftp6": "Accessible FTP IPv6",
    "scan_postgres": "Postgres", "scan6_postgres": "Postgres IPv6", "scan_couchdb": "Couchdb", "scan6_ntp": "NTP-version IPv6", "scan6_ntpmonitor": "NTP-monitor IPv6", "population_bgp": "Accessible BGP",
    "population6_bgp": "Accessible BGP IPv6", "population_http_proxy": "Accessible HTTP Proxy", "population6_http_proxy": "Accessible HTTP Proxy IPv6", "population_msmq": "Accessible MSMQ", "population6_msmq": "Accessible MSMQ IPv6",
    "scan_sip": "Accessible SIP", "scan_slp": "Accessible SLP", "scan6_slp": "Accessible SLP IPv6", "scan_stun": "Accessible STUN", "scan6_stun": "Accessible STUN IPv6", "scan_ws_discovery": "Accessible WS-Discovery",
    "scan6_dns": "DNS Open Resolver IPv6", "scan6_ssl_poodle": "SSL POODLE IPv6", "scan6_snmp": "SNMP IPv6", "scan6_http_vulnerable": "HTTP Vulnerable IPv6"
}

descriptions_db = pd.read_excel(descriptions_path)


def create_workbook(folder):
    """
    Create a new workbook at the location where 'folder' is.
    """
    new_wb = openpyxl.Workbook()
    sheet = new_wb.active
    sheet.title = "Incident Register"
    adv_directory = os.path.abspath(folder)
    register_name = "Incident Register-" + strftime("%d-%m-%Y") + ".xlsx"
    parent_directory = Path(adv_directory).parent
    final_path = parent_directory / register_name
    try:
        new_wb.save(final_path)
    except PermissionError:
        cprint("[!!!] Error: The register file seems to be open.", "red")
        exit(0)
    return final_path


def determine_req(file_name):
    """
    Determine category and requestor
    NB: tuple(category, requestor)
    """
    f_name = file_name.lower()
    if("andromeda" in f_name or "necurs" in f_name or "pykspa" in f_name or
        "suppobox" in f_name or "emotet" in f_name or "bedep" in f_name or
        "nymaim" in f_name or "qsnatch" in f_name or "tinba" in f_name or
         "zeus" in f_name or "ranbyus" in f_name or "matsnu" in f_name or
        "m0yv" in f_name or "gamarue" in f_name):
        category = "Malware"
        requestor = "CERT-BUND"
        req_tuple = (category, requestor)
    elif("bruteforce_global" in f_name ):
        category = "Botnet/DDOS"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("bruteforce_kenya" in f_name ):
        category = "Botnet/DDOS"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("loginattempts" in f_name):
        category = "Botnet/DDOS"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("idsevents_global" in f_name):
        category = "Botnet/DDOS"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("elastic_global" in f_name):
        category = "System Misconfiguration"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("rdp_global" in f_name):
        category = "System Misconfiguration"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)    
    elif("adb_global" in f_name):
        category = "Reconnaissance Attack"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("adb_kenya" in f_name):
        category = "Botnet/DDOS"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("malware_global"  in f_name):
        category = "Malware"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("malware_kenya" in f_name):
        category = "Malware"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("webapp_global" in f_name):
        category = "Web Application"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("webapp_kenya" in f_name):
        category = "Web Application"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("tcpudp_global" in f_name):
        category = "Web Application"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("tcpudp_kenya" in f_name):
        category = "Web Application"
        requestor = "HONEYPOT Advisory"
        req_tuple = (category, requestor)
    elif("phishing" in f_name):
        category = "Phishing"
        requestor = "IntelMQ"
        req_tuple = (category, requestor)
    elif("intel_blacklist" in f_name):
        category = "Blacklist"
        requestor = "IntelMQ"
        req_tuple = (category, requestor)
    elif("intel_darknet" in f_name):
        category = "Darknet"
        requestor = "IntelMQ"
        req_tuple = (category, requestor)
    elif("intel_ssh" in f_name):
        category = "Botnet/DDoS"
        requestor = "IntelMQ"
        req_tuple = (category, requestor)
    elif("ransomware" in f_name):
        category = "Ransomware"
        requestor = "IntelMQ"
        req_tuple = (category, requestor)
        #Brute Force
    elif("honeypot_brute_force" in f_name):
        category = "Brute Force"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #Man-in-the-Middle Attacks
    elif("freak" in f_name or "poodle" in f_name):
        category = "Man-in-the-Middle Attacks"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #IP Spoofing
    elif("ip_spoofer" in f_name):
        category = "IP Spoofing"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #Darknet
    elif("honeypot_darknet" in f_name or "darknet" in f_name):
        category = "Darknet"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #DNS Sinkholing
    elif("microsoft_sinkhole" in f_name ):
        category = "DNS Sinkholing"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #Compromised Website/Domain
    elif("compromised_website" in f_name ):
        category = "Compromised Website/Domain"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #Blacklisted IPs
    elif("blocklist-kenya" in f_name or "blacklist_local" in f_name):
        category = "Blacklisted IPs"
        requestor = "Man-in-the-Middle Attacks"
        req_tuple = (category, requestor)
        #DDoS/Botnet
    elif("honeypot_ddos" in f_name or "botnet_drone" in f_name or "scan_dns" in f_name
         or "mssql" in f_name or "netbios" in f_name or "ubiquiti" in f_name
         or "portmapper" in f_name or "scan_qotd" in f_name or "ssdp" in f_name or "sandbox_url" in f_name):
        category = "Botnet/DDoS"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
        #Reconnaissance Attacks
    elif("honeypot_http_scan" in f_name or "isakmp" in f_name or "mdns" in f_name
         or "memcached" in f_name or "mongodb" in f_name or "ms-rdpeudp" in f_name
         or "nat_pmp" in f_name or "scan_ntp-" in f_name or "scan_ntpmonitor" in f_name
         or "scan_redis" in f_name or "scan_elasticsearch" in f_name or "db2" in f_name
         or "ipp" in f_name or "scan_ntp-" in f_name or "tftp" in f_name
         or "chargen" in f_name or "scan_cwmp" in f_name or "scan_radmin" in f_name
         or "rdp" in f_name or "smb" in f_name or "telnet" in f_name or "device_id" in f_name):
        category = "Reconnaissance Attacks"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
    else:
        category = "System Misconfiguration"
        requestor = "Shadow Server"
        req_tuple = (category, requestor)
    return req_tuple


def determine_description(f_path):
    """
    Determine what should be written under
    detailed description.
    """
    for key in advisory_dict:
        if key in os.path.basename(f_path).lower():
            # print("Here: ")
            desc = descriptions_db.loc[descriptions_db["Advisory"] ==
                                       advisory_dict[key], "Detail Description"].iloc[0]
    return desc


def determine_local_or_global(folder_path):
    """
    Determine if an advisory is local or global.
    """
    folder_name = os.path.basename(folder_path).lower()
    is_global = (
        "_global" in folder_name or
        "ransomware" in folder_name
    )
    if(is_global):
        location = "Global"
    else:
        location = "Local"
    return location


def get_isps(f_path):
    """
    Get a list of isps represented in a folder.
    """
    if("_global" in os.path.basename(f_path)):
        isps_in_folder = get_countries_in_folder(f_path)
    else:
        isps_in_folder = get_isps_in_folder(f_path)
        if("accesskenya" in isps_in_folder and
            "is" in isps_in_folder
           ):
            isps_in_folder.remove("is")
        if("safaricom(1)" in isps_in_folder and
            "safaricom(2)" in isps_in_folder
           ):
            isps_in_folder.remove("safaricom(2)")
    return isps_in_folder


def get_subject(isp, folder_path):
    """
    Determine the subject of the entry in the register.
    """
    is_international = (
        "_global" in os.path.basename(folder_path) or
        "ransomware" in os.path.basename(folder_path) or
        "scan_mysql" in os.path.basename(folder_path)
    )
    if(is_international):
        subject = "Vulnerable hosts on " + titlecase(isp) + " network"
    else:
        subject = "Vulnerable hosts on " + isp_dict[isp]
    return subject


def get_files(f_path):
    """
    Get a list of filenames of files in a folder.
    """
    files_in_folder = get_file_names(f_path)
    return files_in_folder


def populate_register(workbook_path, folder_path, ref_no, count=1):
    """
    Add entries to the Excel workbook
    """
    workbook = openpyxl.load_workbook(workbook_path)
    location = determine_local_or_global(folder_path)
    isps_in_folder = get_isps(folder_path)
    files_in_folder = get_files(folder_path)
    sheet = workbook["Incident Register"]
    index = 0
    for isp in isps_in_folder:  # or countries
        requestor_tuple = determine_req(files_in_folder[index])
        isp = get_const_isp(isp)
        sheet.cell(column=1, row=count).value = ref_no
        sheet.cell(column=2, row=count).value = strftime("%d.%m.%y")
        sheet.cell(column=4, row=count).value = requestor_tuple[0]
        sheet.cell(column=5, row=count).value = get_subject(isp, folder_path)
        sheet.cell(column=6, row=count).value = determine_description(
            folder_path)
        sheet.cell(column=7, row=count).value = requestor_tuple[1]
        sheet.cell(column=8, row=count).value = location
        sheet.cell(column=9, row=count).value = status
        sheet.cell(column=10, row=count).value = "Resolved"
        sheet.cell(column=11, row=count).value = strftime("%d.%m.%y")
        count = count + 1
        ref_no = ref_no + 1
    workbook.save(workbook_path)
    cprint("[!] Info: Last ref number: " + str(ref_no - 1) + "\n", "green")
    return count


def main():
    if(len(sys.argv) >= 2):
        new_register = create_workbook(sys.argv[1])
        row_count = 1
        for folder in sys.argv[1:]:
            no = False
            start_ref = 0
            while(not(no)):
                try:
                    start_ref = int(input("Enter starting reference number: "))
                except ValueError:
                    cprint("[!!] Warning: You didn\'t enter a digit.", "yellow")
                else:
                    no = True
            cprint("\n[!] Working with folder: " +
                   str(os.path.basename(folder)) + ".\n", "green")
            try:
                row_count = populate_register(new_register, os.path.abspath(
                    folder), start_ref, row_count)
            except NotADirectoryError:
                cprint("[!!!] Error " + os.path.basename(folder) +
                       " is not a directory.", "red")
                exit(0)
    else:
        cprint("[!!!] Error: No folder was specified.", "red")
        exit(0)


if __name__ == "__main__":
    main()